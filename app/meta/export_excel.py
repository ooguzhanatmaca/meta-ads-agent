"""Export existing read-only Meta reports to a formatted Excel workbook."""

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.meta.client import (
    MetaAPIError,
    get_account_insights_for_period,
    get_performance_report,
)
from app.meta.compare_periods import (
    METRIC_LABELS,
    build_default_periods,
    calculate_period_metrics,
    compare_metrics,
)
from app.meta.performance_report import calculate_report_rows
from app.rules.performance_rules import evaluate_ads


CURRENCY_FORMAT = '₺#,##0.00'
NUMBER_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00%'
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CRITICAL_FILL = PatternFill("solid", fgColor="F4CCCC")
HIGH_FILL = PatternFill("solid", fgColor="FCE5CD")
BAD_ROAS_FILL = PatternFill("solid", fgColor="F4CCCC")
GOOD_ROAS_FILL = PatternFill("solid", fgColor="D9EAD3")


def collect_report_data() -> dict[str, Any]:
    """Collect all report data through the existing read-only functions."""
    comparisons = []
    period_metrics = []
    for period in build_default_periods():
        current = calculate_period_metrics(
            get_account_insights_for_period(
                str(period.current_since), str(period.current_until)
            )
        )
        previous = calculate_period_metrics(
            get_account_insights_for_period(
                str(period.previous_since), str(period.previous_until)
            )
        )
        period_metrics.append((current, previous))
        comparisons.append((period, compare_metrics(current, previous)))

    campaigns = calculate_report_rows(
        get_performance_report("campaign", "last_7d")
    )
    adsets = calculate_report_rows(get_performance_report("adset", "last_7d"))
    ads = calculate_report_rows(get_performance_report("ad", "last_7d"))
    return {
        "today": period_metrics[0][0],
        "comparisons": comparisons,
        "campaigns": campaigns,
        "adsets": adsets,
        "ads": ads,
        "recommendations": evaluate_ads(ads),
    }


def _style_header(worksheet, row: int, columns: int) -> None:
    for cell in worksheet[row][:columns]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _finish_sheet(worksheet, max_width: int = 42) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _style_header(worksheet, 1, worksheet.max_column)
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[letter].width = min(
            max(length + 2, 12), max_width
        )


def _add_table(worksheet, name: str) -> None:
    if worksheet.max_row < 2:
        return
    reference = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _format_columns(worksheet, formats: dict[str, str]) -> None:
    headers = {cell.value: cell.column for cell in worksheet[1]}
    for header, number_format in formats.items():
        column = headers.get(header)
        if column is None:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row, column).number_format = number_format


def _add_roas_formatting(worksheet) -> None:
    headers = {cell.value: cell.column for cell in worksheet[1]}
    column = headers.get("ROAS")
    if column is None or worksheet.max_row < 2:
        return
    letter = get_column_letter(column)
    cell_range = f"{letter}2:{letter}{worksheet.max_row}"
    worksheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["1.5"], fill=BAD_ROAS_FILL),
    )
    worksheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["2.5"],
            fill=GOOD_ROAS_FILL,
        ),
    )


def _priority_fill(priority: str):
    if priority == "critical":
        return CRITICAL_FILL
    if priority == "high":
        return HIGH_FILL
    return None


def _write_executive_summary(workbook: Workbook, data: dict[str, Any]) -> None:
    worksheet = workbook.active
    worksheet.title = "Yönetici Özeti"
    worksheet.append(["Metrik", "Değer"])
    today = data["today"]
    for row in (
        ("Harcama", today["spend"]),
        ("Satın alma", today["purchases"]),
        ("Satış değeri", today["purchase_value"]),
        ("CPA", today["cpa"]),
        ("ROAS", today["roas"]),
        ("CTR", today["ctr"] / 100),
        ("CPC", today["cpc"]),
        ("CPM", today["cpm"]),
        ("Frekans", today["frequency"]),
    ):
        worksheet.append(row)

    action_row = worksheet.max_row + 2
    for column, value in enumerate(
        ("Öncelikli aksiyonlar", "Reklam", "Öneri", "Gerekçe", "Öncelik"),
        start=1,
    ):
        worksheet.cell(action_row, column, value)
    _style_header(worksheet, action_row, 5)
    for item in data["recommendations"]:
        worksheet.append(
            [
                "Aksiyon",
                item["name"],
                item["recommendation"],
                item["reason"],
                item["priority"],
            ]
        )
        fill = _priority_fill(item["priority"])
        if fill:
            for cell in worksheet[worksheet.max_row]:
                cell.fill = fill

    _style_header(worksheet, 1, 2)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:B10"
    for column, width in {
        "A": 24,
        "B": 22,
        "C": 34,
        "D": 62,
        "E": 12,
    }.items():
        worksheet.column_dimensions[column].width = width
    for row in (2, 4, 5, 8, 9):
        worksheet.cell(row, 2).number_format = CURRENCY_FORMAT
    worksheet.cell(6, 2).number_format = NUMBER_FORMAT
    worksheet.cell(7, 2).number_format = PERCENT_FORMAT
    worksheet.cell(10, 2).number_format = NUMBER_FORMAT


def _write_comparisons(workbook: Workbook, data: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("Dönem Karşılaştırması")
    worksheet.append(
        [
            "Karşılaştırma",
            "Metrik",
            "Mevcut değer",
            "Önceki değer",
            "Yüzdesel değişim",
            "Değerlendirme",
        ]
    )
    currency_metrics = {"spend", "purchase_value", "cpa", "cpc", "cpm"}
    for period, comparisons in data["comparisons"]:
        for item in comparisons:
            current = item["current"]
            previous = item["previous"]
            if item["metric"] == "ctr":
                current /= 100
                previous /= 100
            worksheet.append(
                [
                    period.title,
                    METRIC_LABELS[item["metric"]],
                    current,
                    previous,
                    item["change"] / 100,
                    item["result"],
                ]
            )
            row = worksheet.max_row
            worksheet.cell(row, 5).number_format = PERCENT_FORMAT
            if item["metric"] in currency_metrics:
                worksheet.cell(row, 3).number_format = CURRENCY_FORMAT
                worksheet.cell(row, 4).number_format = CURRENCY_FORMAT
            elif item["metric"] == "ctr":
                worksheet.cell(row, 3).number_format = PERCENT_FORMAT
                worksheet.cell(row, 4).number_format = PERCENT_FORMAT
            else:
                worksheet.cell(row, 3).number_format = NUMBER_FORMAT
                worksheet.cell(row, 4).number_format = NUMBER_FORMAT
    _finish_sheet(worksheet)
    _add_table(worksheet, "PeriodComparisons")


PERFORMANCE_HEADERS = [
    "Durum",
    "Harcama",
    "Gösterim",
    "Erişim",
    "Tıklama",
    "CTR",
    "CPC",
    "CPM",
    "Frekans",
    "Satın alma",
    "Satış değeri",
    "CPA",
    "ROAS",
]


def _performance_values(row: dict[str, Any]) -> list[Any]:
    return [
        row["status"],
        row["spend"],
        row["impressions"],
        row["reach"],
        row["clicks"],
        row["ctr"] / 100,
        row["cpc"],
        row["cpm"],
        row["frequency"],
        row["purchases"],
        row["purchase_value"],
        row["cpa"],
        row["roas"],
    ]


def _write_performance_sheet(
    workbook: Workbook,
    title: str,
    identity_headers: list[str],
    identity_keys: list[str],
    rows: list[dict[str, Any]],
    table_name: str,
) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(identity_headers + PERFORMANCE_HEADERS)
    for row in rows:
        worksheet.append(
            [row.get(key, "-") for key in identity_keys]
            + _performance_values(row)
        )
    _finish_sheet(worksheet)
    _add_table(worksheet, table_name)
    _format_columns(
        worksheet,
        {
            "Harcama": CURRENCY_FORMAT,
            "CTR": PERCENT_FORMAT,
            "CPC": CURRENCY_FORMAT,
            "CPM": CURRENCY_FORMAT,
            "Satış değeri": CURRENCY_FORMAT,
            "CPA": CURRENCY_FORMAT,
            "ROAS": NUMBER_FORMAT,
            "Frekans": NUMBER_FORMAT,
        },
    )
    _add_roas_formatting(worksheet)


def _write_recommendations(workbook: Workbook, data: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("Öneriler")
    worksheet.append(
        [
            "Reklam adı",
            "Harcama",
            "Satın alma",
            "CPA",
            "ROAS",
            "Frekans",
            "Öneri",
            "Gerekçe",
            "Öncelik",
        ]
    )
    for item in data["recommendations"]:
        worksheet.append(
            [
                item["name"],
                item["spend"],
                item["purchases"],
                item["cpa"],
                item["roas"],
                item["frequency"],
                item["recommendation"],
                item["reason"],
                item["priority"],
            ]
        )
        fill = _priority_fill(item["priority"])
        if fill:
            for cell in worksheet[worksheet.max_row]:
                cell.fill = fill
    _finish_sheet(worksheet, max_width=60)
    _add_table(worksheet, "Recommendations")
    _format_columns(
        worksheet,
        {
            "Harcama": CURRENCY_FORMAT,
            "CPA": CURRENCY_FORMAT,
            "ROAS": NUMBER_FORMAT,
            "Frekans": NUMBER_FORMAT,
        },
    )
    _add_roas_formatting(worksheet)


def build_workbook(data: dict[str, Any]) -> Workbook:
    """Build the six-sheet Meta Ads workbook."""
    workbook = Workbook()
    _write_executive_summary(workbook, data)
    _write_comparisons(workbook, data)
    _write_performance_sheet(
        workbook,
        "Kampanyalar",
        ["Kampanya ID", "Kampanya adı"],
        ["id", "name"],
        data["campaigns"],
        "Campaigns",
    )
    _write_performance_sheet(
        workbook,
        "Reklam Setleri",
        ["Reklam seti ID", "Reklam seti adı", "Kampanya adı"],
        ["id", "name", "campaign_name"],
        data["adsets"],
        "AdSets",
    )
    _write_performance_sheet(
        workbook,
        "Reklamlar",
        ["Reklam ID", "Reklam adı", "Reklam seti adı", "Kampanya adı"],
        ["id", "name", "adset_name", "campaign_name"],
        data["ads"],
        "Ads",
    )
    _write_recommendations(workbook, data)
    return workbook


def export_excel(
    output_dir: Path | str = "reports",
    now: datetime | None = None,
) -> Path:
    """Collect data and save a timestamped Excel workbook."""
    timestamp = now or datetime.now()
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    output_path = directory / f"meta_ads_report_{timestamp:%Y-%m-%d_%H-%M}.xlsx"
    build_workbook(collect_report_data()).save(output_path)
    return output_path


def main() -> int:
    try:
        output_path = export_excel()
        print(f"Excel raporu oluşturuldu: {output_path}")
    except MetaAPIError as error:
        print(f"Excel raporu oluşturulamadı: {error}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
