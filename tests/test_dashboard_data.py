from datetime import date

from openpyxl import load_workbook

from app.dashboard_data import (
    dashboard_excel_bytes,
    load_dashboard_data,
    previous_period,
    recommendations_dataframe,
    rows_to_dataframe,
)


def test_previous_period_has_same_inclusive_length() -> None:
    start, end = previous_period(date(2026, 6, 5), date(2026, 6, 11))

    assert start == date(2026, 5, 29)
    assert end == date(2026, 6, 4)


def test_load_dashboard_data_reuses_existing_report_functions(monkeypatch) -> None:
    insight_calls = []
    report_calls = []

    def mock_insights(since, until):
        insight_calls.append((since, until))
        return {
            "spend": "100",
            "impressions": "1000",
            "reach": "500",
            "clicks": "20",
        }

    def mock_report(level, since, until):
        report_calls.append((level, since, until))
        return []

    monkeypatch.setattr(
        "app.dashboard_data.get_account_insights_for_period",
        mock_insights,
    )
    monkeypatch.setattr(
        "app.dashboard_data.get_performance_report_for_period",
        mock_report,
    )

    data = load_dashboard_data(date(2026, 6, 5), date(2026, 6, 11))

    assert insight_calls == [
        ("2026-06-05", "2026-06-11"),
        ("2026-05-29", "2026-06-04"),
    ]
    assert report_calls == [
        ("campaign", "2026-06-05", "2026-06-11"),
        ("adset", "2026-06-05", "2026-06-11"),
        ("ad", "2026-06-05", "2026-06-11"),
    ]
    assert data["current"]["frequency"] == 2
    assert data["recommendations"] == []


def test_load_dashboard_data_rejects_invalid_range() -> None:
    try:
        load_dashboard_data(date(2026, 6, 11), date(2026, 6, 5))
    except ValueError as error:
        assert "Başlangıç" in str(error)
    else:
        raise AssertionError("Geçersiz tarih aralığı reddedilmedi.")


def test_dataframe_helpers_handle_empty_and_populated_rows() -> None:
    assert rows_to_dataframe([], ["id", "name"]).empty
    frame = rows_to_dataframe(
        [{"id": "1", "name": "Ad", "spend": 100}],
        ["id", "name", "spend"],
    )
    assert list(frame.columns) == ["ID", "Ad", "Harcama"]
    recommendations = recommendations_dataframe(
        [
            {
                "name": "Ad",
                "spend": 100,
                "purchases": 0,
                "cpa": 0,
                "roas": 0,
                "frequency": 2,
                "recommendation": "Kapatılmaya aday",
                "reason": "Test",
                "priority": "critical",
            }
        ]
    )
    assert recommendations.iloc[0]["Öncelik"] == "critical"


def test_dashboard_excel_bytes_builds_workbook_without_api_calls() -> None:
    metrics = {
        "spend": 0,
        "impressions": 0,
        "reach": 0,
        "clicks": 0,
        "ctr": 0,
        "cpc": 0,
        "cpm": 0,
        "frequency": 0,
        "purchases": 0,
        "purchase_value": 0,
        "cpa": 0,
        "roas": 0,
    }
    content = dashboard_excel_bytes(
        {
            "current": metrics,
            "campaigns": [],
            "adsets": [],
            "ads": [],
            "recommendations": [],
        }
    )

    workbook = load_workbook(filename=__import__("io").BytesIO(content))
    assert workbook.sheetnames[0] == "Yönetici Özeti"
