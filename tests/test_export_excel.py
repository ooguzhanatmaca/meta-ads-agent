from datetime import date, datetime

from openpyxl import load_workbook

from app.meta.compare_periods import build_default_periods
from app.meta.export_excel import (
    CURRENCY_FORMAT,
    build_workbook,
    collect_report_data,
    export_excel,
)


def sample_row(
    name="Test",
    roas=3,
    campaign_name="Campaign",
    adset_name="Ad Set",
):
    return {
        "id": "123",
        "name": name,
        "campaign_name": campaign_name,
        "adset_name": adset_name,
        "status": "ACTIVE",
        "spend": 1000,
        "impressions": 10000,
        "reach": 5000,
        "clicks": 100,
        "ctr": 1,
        "cpc": 10,
        "cpm": 100,
        "frequency": 2,
        "purchases": 5,
        "purchase_value": 3000,
        "cpa": 200,
        "roas": roas,
    }


def sample_data():
    metrics = {
        "spend": 100,
        "impressions": 1000,
        "reach": 500,
        "clicks": 20,
        "ctr": 2,
        "cpc": 5,
        "cpm": 100,
        "frequency": 2,
        "purchases": 2,
        "purchase_value": 400,
        "cpa": 50,
        "roas": 4,
    }
    comparison = [
        {
            "metric": "roas",
            "current": 4,
            "previous": 2,
            "change": 100,
            "direction": "arttı",
            "result": "iyileşti",
        }
    ]
    recommendation = {
        "name": "Test Ad",
        "spend": 1000,
        "purchases": 0,
        "cpa": 0,
        "roas": 0,
        "frequency": 2,
        "recommendation": "Kapatılmaya aday",
        "reason": "Satın alma yok.",
        "priority": "critical",
    }
    return {
        "today": metrics,
        "comparisons": [
            (build_default_periods(date(2026, 6, 11))[0], comparison)
        ],
        "campaigns": [sample_row("Campaign")],
        "adsets": [sample_row("Ad Set")],
        "ads": [sample_row("Ad")],
        "recommendations": [recommendation],
    }


def test_build_workbook_contains_required_sheets_and_formatting() -> None:
    workbook = build_workbook(sample_data())

    assert workbook.sheetnames == [
        "Yönetici Özeti",
        "Dönem Karşılaştırması",
        "Kampanyalar",
        "Reklam Setleri",
        "Reklamlar",
        "Öneriler",
    ]
    campaigns = workbook["Kampanyalar"]
    assert campaigns.freeze_panes == "A2"
    assert campaigns.auto_filter.ref
    assert campaigns["B2"].value == "Campaign"
    assert campaigns["D2"].number_format == CURRENCY_FORMAT
    rule_count = sum(
        len(item.rules) for item in campaigns.conditional_formatting
    )
    assert rule_count == 2
    assert workbook["Reklam Setleri"]["C2"].value == "Campaign"
    assert workbook["Reklamlar"]["C2"].value == "Ad Set"
    assert workbook["Öneriler"]["A2"].fill.fgColor.rgb.endswith("F4CCCC")


def test_collect_report_data_uses_mocked_existing_functions(monkeypatch) -> None:
    periods = build_default_periods(date(2026, 6, 11))
    monkeypatch.setattr(
        "app.meta.export_excel.build_default_periods", lambda: periods
    )
    insight_calls = []
    report_calls = []

    def mock_insights(since, until):
        insight_calls.append((since, until))
        return {"spend": "100", "impressions": "1000", "reach": "500"}

    def mock_report(level, preset):
        report_calls.append((level, preset))
        return []

    monkeypatch.setattr(
        "app.meta.export_excel.get_account_insights_for_period",
        mock_insights,
    )
    monkeypatch.setattr(
        "app.meta.export_excel.get_performance_report",
        mock_report,
    )

    data = collect_report_data()

    assert len(insight_calls) == 6
    assert report_calls == [
        ("campaign", "last_7d"),
        ("adset", "last_7d"),
        ("ad", "last_7d"),
    ]
    assert data["recommendations"] == []


def test_export_excel_creates_timestamped_file(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(
        "app.meta.export_excel.collect_report_data",
        sample_data,
    )

    output = export_excel(tmp_path, datetime(2026, 6, 11, 14, 35))

    assert output.name == "meta_ads_report_2026-06-11_14-35.xlsx"
    assert output.exists()
    workbook = load_workbook(output)
    assert "Yönetici Özeti" in workbook.sheetnames
    assert "test-token" not in str(output)
